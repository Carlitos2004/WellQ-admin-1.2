"""
routers/clinics.py — MÓDULO: GESTIÓN DE CLÍNICAS
Endpoints 14 al 25 conectados a Neon (PostgreSQL)
"""

from fastapi import APIRouter, Depends, HTTPException, Path, Body, status, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc, asc
from datetime import datetime
import uuid

from app.models_db import (
    Clinic, ClinicPlan, ClinicUsageMetric, 
    Invoice, Notification, Job
)
from app.db.neon import get_db

router = APIRouter(prefix="/api/clinics", tags=["Gestión de Clínicas"])


# ─────────────────────────────────────────────────────────────────────────────
# 14. GET /clinics — Listar clínicas con filtros y paginación
# ─────────────────────────────────────────────────────────────────────────────
@router.get("", summary="Listar clínicas con filtros y paginación")
async def list_clinics(
    search: str | None = None,
    tier: str | None = None,
    status_param: str | None = Query(None, alias="status"),
    page: int = 1,
    page_size: int = 20,
    sort_by: str = "name",
    sort_order: str = "asc",
    db: AsyncSession = Depends(get_db)
):
    # Construir query base
    query = select(Clinic)
    
    # Filtros
    if search:
        query = query.where(Clinic.name.ilike(f"%{search}%"))
    if tier:
        query = query.where(Clinic.tier == tier)
    if status_param:
        query = query.where(Clinic.status == status_param)

    # Calcular el total para la paginación
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Ordenamiento dinámico
    order_col = getattr(Clinic, sort_by, Clinic.name)
    if sort_order == "desc":
        query = query.order_by(desc(order_col))
    else:
        query = query.order_by(asc(order_col))

    # Paginación
    query = query.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(query)
    clinics = result.scalars().all()

    # Mapear respuesta
    data = []
    for c in clinics:
        data.append({
            "_id": str(c.id),
            "clinic_id": c.clinic_id,
            "name": c.name,
            "tier": c.tier,
            "status": c.status,
            "contact": {
                "phone": getattr(c, 'contact_phone', None),
                "email": getattr(c, 'contact_email', None)
            },
            "patient_count": getattr(c, 'patients_used', 0),
            "patientsUsed": getattr(c, 'patients_used', 0),
            "patientsLimit": getattr(c, 'patients_limit', 500),
            "healthScore": getattr(c, 'health_score', 100),
            "lastLogin": c.last_login.isoformat() if getattr(c, 'last_login', None) else None
        })

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "data": data
    }


# ─────────────────────────────────────────────────────────────────────────────
# 15. POST /clinics — Registro de una nueva clínica en el sistema
# FIX: Ahora captura TODOS los datos del modal de React
# ─────────────────────────────────────────────────────────────────────────────
@router.post("", summary="Registro de una nueva clínica", status_code=status.HTTP_201_CREATED)
async def create_clinic(body: dict = Body(...), db: AsyncSession = Depends(get_db)):
    # Generar un ID único para la nueva clínica
    new_clinic_id = f"CL-{uuid.uuid4().hex[:6].upper()}"
    
    # Capturamos todos los campos enviados desde React, si no vienen, dejamos defaults
    new_clinic = Clinic(
        clinic_id=new_clinic_id,
        name=body.get("name", "Nueva Clínica"),
        tier=body.get("tier", "smb"),
        status=body.get("status", "active"), # Ahora sí respeta el "active" de React
        patients_limit=body.get("patients_limit", 500),
        mrr=body.get("mrr", 0.0),
        location=body.get("location"),
        contact_name=body.get("contact_name"),
        contact_email=body.get("contact_email"),
        contact_phone=body.get("contact_phone"),
        company_name=body.get("company_name"),
        tax_id=body.get("tax_id"),
        billing_email=body.get("billing_email"),
        address=body.get("address"),
        internal_notes=body.get("internal_notes")
    )
    
    db.add(new_clinic)
    await db.commit()
    await db.refresh(new_clinic)

    return {
        "status": "success",
        "message": "Clínica registrada correctamente",
        "data": {
            "_id": str(new_clinic.id),
            "clinic_id": new_clinic.clinic_id,
            "name": new_clinic.name,
            "status": new_clinic.status,
            "tier": new_clinic.tier
        }
    }


# ─────────────────────────────────────────────────────────────────────────────
# 24. POST /clinics/bulk/email — Envío de comunicaciones masivas
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/bulk/email", summary="Envío de comunicaciones masivas a clínicas")
async def bulk_email(body: dict = Body(...), db: AsyncSession = Depends(get_db)):
    clinic_ids = body.get('clinic_ids', [])
    subject = body.get("subject", "Actualización importante")
    
    notif = Notification(
        notification_id=f"notif-{uuid.uuid4().hex[:8]}",
        title=subject,
        message=f"Mensaje masivo enviado a {len(clinic_ids)} clínicas.",
        channel="email",
        status="pending",
        recipient_clinic_id="multiple",
        sent_by="admin_system"
    )
    
    db.add(notif)
    await db.commit()

    return {
        "status": "success",
        "message": f"Correos encolados para {len(clinic_ids)} clínicas.",
        "subject": subject
    }


# ─────────────────────────────────────────────────────────────────────────────
# 25. GET /clinics/export — Exportación real de clínicas como XLSX con estilos
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export", summary="Exportación de lista de clínicas en XLSX con colores")
async def export_clinics(
    status_param: str | None = Query(None, alias="status"),
    tier: str | None = Query(None),
    db: AsyncSession = Depends(get_db)
):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    C_DARK    = "1A1A2E"
    C_WHITE   = "FFFFFF"
    C_TEAL    = "0D9488"
    C_ALT_ROW = "E6FAFA"

    STATUS_COLORS = {
        "active":    ("D1FAE5", "065F46"),
        "warning":   ("FEF3C7", "92400E"),
        "critical":  ("FEE2E2", "991B1B"),
        "churned":   ("F3F4F6", "374151"),
        "trial":     ("EDE9FE", "5B21B6"),
        "onboarding":("DBEAFE", "1E40AF"),
    }
    TIER_COLORS = {
        "enterprise": ("1A1A2E", "FFFFFF"),
        "smb":        ("E0F2FE", "075985"),
        "trial":      ("F3E8FF", "6B21A8"),
    }

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style='thin', color='E5E7EB')
        return Border(left=s, right=s, top=s, bottom=s)

    query = select(Clinic)
    if status_param:
        query = query.where(Clinic.status == status_param)
    if tier:
        query = query.where(Clinic.tier == tier)
    query = query.order_by(asc(Clinic.name))
    result = await db.execute(query)
    clinics = result.scalars().all()

    data = [{
        "clinic_id": c.clinic_id,
        "name": c.name,
        "tier": c.tier,
        "status": c.status,
        "patientsUsed": getattr(c, 'patients_used', 0),
        "patientsLimit": getattr(c, 'patients_limit', 0),
        "healthScore": getattr(c, 'health_score', 0),
        "mrr": getattr(c, 'mrr', 0),
        "lastLogin": c.last_login.isoformat() if getattr(c, 'last_login', None) else "",
        "location": getattr(c, 'location', ""),
    } for c in clinics]

    wb = Workbook()
    ws = wb.active
    ws.title = "Clínicas"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = f"WellQ — Reporte de Clínicas   |   {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC"
    t.font = Font(name="Arial", bold=True, size=13, color=C_WHITE)
    t.fill = fill(C_DARK)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    headers    = ["ID", "Nombre Clínica", "Tier", "Estado", "Pacientes", "Límite", "Uso %", "Health", "MRR (USD)", "Último Login", "Ciudad"]
    col_widths = [12,   28,               13,     12,       11,          10,       9,       10,       13,          22,             16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        cell.fill      = fill(C_TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, c in enumerate(data, 3):
        bg   = C_ALT_ROW if ri % 2 == 0 else C_WHITE
        uso  = round((c["patientsUsed"] / c["patientsLimit"]) * 100, 1) if c["patientsLimit"] > 0 else 0
        login_str = ""
        if c["lastLogin"]:
            try: login_str = datetime.fromisoformat(c["lastLogin"].replace("Z","")).strftime("%d/%m/%Y %H:%M")
            except: login_str = c["lastLogin"]

        vals = [c["clinic_id"], c["name"], c["tier"].upper(), c["status"].capitalize(),
                c["patientsUsed"], c["patientsLimit"], uso / 100, c["healthScore"],
                c["mrr"], login_str, c["location"]]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name="Arial", size=10, color="1F2937")
            cell.fill      = fill(bg)
            cell.border    = border()
            cell.alignment = Alignment(vertical="center")

        tbg, tfg = TIER_COLORS.get(c["tier"].lower(), ("E5E7EB", "374151"))
        tc = ws.cell(row=ri, column=3)
        tc.fill = fill(tbg); tc.font = Font(name="Arial", size=10, bold=True, color=tfg)
        tc.alignment = Alignment(horizontal="center", vertical="center")

        sbg, sfg = STATUS_COLORS.get(c["status"].lower(), ("F3F4F6", "374151"))
        sc = ws.cell(row=ri, column=4)
        sc.fill = fill(sbg); sc.font = Font(name="Arial", size=10, bold=True, color=sfg)
        sc.alignment = Alignment(horizontal="center", vertical="center")

        pc = ws.cell(row=ri, column=7)
        pc.number_format = "0.0%"
        pc.alignment = Alignment(horizontal="center", vertical="center")
        if uso >= 90:   pc.fill = fill("FEE2E2"); pc.font = Font(name="Arial", size=10, bold=True, color="991B1B")
        elif uso >= 70: pc.fill = fill("FEF3C7"); pc.font = Font(name="Arial", size=10, bold=True, color="92400E")
        else:           pc.fill = fill("D1FAE5"); pc.font = Font(name="Arial", size=10, color="065F46")

        hs = c["healthScore"]
        hc = ws.cell(row=ri, column=8)
        hbg, hfg = ("F0FDF4","065F46") if hs>=80 else ("FFFBEB","92400E") if hs>=60 else ("FFF1F2","991B1B") if hs>0 else ("F9FAFB","9CA3AF")
        hc.fill = fill(hbg); hc.font = Font(name="Arial", size=10, bold=True, color=hfg)
        hc.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=ri, column=9).number_format = '"$"#,##0.00'
        ws.cell(row=ri, column=9).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=ri, column=10).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"clinicas_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ─────────────────────────────────────────────────────────────────────────────
# 16. GET /clinics/{clinic_id} — Obtener detalle de una clínica
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/{clinic_id}", summary="Obtener detalle de una clínica")
async def get_clinic(clinic_id: str = Path(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Clinic).where(Clinic.clinic_id == clinic_id))
    clinic = result.scalar_one_or_none()
    
    if not clinic:
        raise HTTPException(status_code=404, detail="Clínica no encontrada")

    return {
        "_id": str(clinic.id),
        "clinic_id": clinic.clinic_id,
        "name": clinic.name,
        "status": clinic.status,
        "tier": clinic.tier,
        "internal_notes": getattr(clinic, 'internal_notes', None),
        "created_at": clinic.created_at.isoformat() if clinic.created_at else None
    }


# ─────────────────────────────────────────────────────────────────────────────
# 17. PATCH /clinics/{clinic_id} — Actualizar campos de una clínica
# ─────────────────────────────────────────────────────────────────────────────
@router.patch("/{clinic_id}", summary="Actualizar campos de una clínica")
async def update_clinic(
    clinic_id: str = Path(...),
    updates: dict = Body(...),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Clinic).where(Clinic.clinic_id == clinic_id))
    clinic = result.scalar_one_or_none()

    if not clinic:
        raise HTTPException(status_code=404, detail="Clínica no encontrada")

    field_map = {
        "name":   "name",
        "tier":   "tier",
        "status": "status",
    }

    updated = []
    for key, value in updates.items():
        model_field = field_map.get(key, key)
        if hasattr(clinic, model_field):
            setattr(clinic, model_field, value)
            updated.append(key)

    clinic.updated_at = datetime.utcnow()
    await db.commit()

    return {
        "status": "success",
        "message": f"Clínica {clinic_id} actualizada correctamente",
        "updated_fields": updated
    }


# ─────────────────────────────────────────────────────────────────────────────
# 17b. DELETE /clinics/{clinic_id} — Eliminar clínica
# ─────────────────────────────────────────────────────────────────────────────
@router.delete("/{clinic_id}", summary="Eliminar clínica del sistema")
async def delete_clinic(
    clinic_id: str = Path(...),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Clinic).where(Clinic.clinic_id == clinic_id))
    clinic = result.scalar_one_or_none()

    if not clinic:
        raise HTTPException(status_code=404, detail="Clínica no encontrada")

    await db.delete(clinic)
    await db.commit()

    return {
        "status": "success",
        "message": f"Clínica {clinic_id} eliminada correctamente"
    }


# ─────────────────────────────────────────────────────────────────────────────
# 18. GET /clinics/{clinic_id}/contact — Info de contacto y facturación
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/{clinic_id}/contact", summary="Información de contacto y facturación")
async def get_clinic_contact(clinic_id: str = Path(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Clinic).where(Clinic.clinic_id == clinic_id))
    clinic = result.scalar_one_or_none()
    
    if not clinic:
        raise HTTPException(status_code=404, detail="Clínica no encontrada")

    return {
        "clinic_id": clinic.clinic_id,
        "contact_info": {
            "primary_name": getattr(clinic, 'contact_name', None),
            "primary_email": getattr(clinic, 'contact_email', None),
            "primary_phone": getattr(clinic, 'contact_phone', None)
        },
        "billing_info": {
            "company_name": getattr(clinic, 'company_name', None),
            "tax_id": getattr(clinic, 'tax_id', None),
            "billing_email": getattr(clinic, 'billing_email', None),
            "address": getattr(clinic, 'address', None)
        }
    }


# ─────────────────────────────────────────────────────────────────────────────
# 19. GET /clinics/{clinic_id}/subscription — Detalles del plan
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/{clinic_id}/subscription", summary="Detalles del plan de suscripción")
async def get_clinic_subscription(clinic_id: str = Path(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(ClinicPlan)
        .where(ClinicPlan.clinic_id == clinic_id)
        .order_by(desc(ClinicPlan.effective_from))
        .limit(1)
    )
    plan_assignment = result.scalar_one_or_none()
    
    if not plan_assignment:
        raise HTTPException(status_code=404, detail="Sin suscripción activa")

    import json
    plan_data = json.loads(plan_assignment.plan_snapshot) if plan_assignment.plan_snapshot else {}

    return {
        "clinic_id": clinic_id,
        "subscription": {
            "plan_name": plan_data.get("name", "Desconocido"),
            "status": "active",
            "mrr_value": plan_data.get("monthlyPrice", 0.0),
            "currency": plan_data.get("currency", "USD"),
            "started_at": plan_assignment.effective_from.isoformat() if plan_assignment.effective_from else None,
            "renews_at": plan_assignment.effective_to.isoformat() if plan_assignment.effective_to else None,
            "features_enabled": ["custom_branding", "api_access", "priority_support"]
        }
    }


# ─────────────────────────────────────────────────────────────────────────────
# 20. GET /clinics/{clinic_id}/usage — Estadísticas de uso
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/{clinic_id}/usage", summary="Estadísticas de uso de la plataforma")
async def get_clinic_usage(clinic_id: str = Path(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(ClinicUsageMetric)
        .where(ClinicUsageMetric.clinic_id == clinic_id)
        .order_by(desc(ClinicUsageMetric.recorded_at))
        .limit(1)
    )
    usage = result.scalar_one_or_none()
    
    if not usage:
        raise HTTPException(status_code=404, detail="Métricas no encontradas para esta clínica")

    return {
        "clinic_id": clinic_id,
        "period": usage.period,
        "metrics": {
            "active_clinicians": usage.active_clinicians,
            "patient_sessions_completed": usage.patient_sessions_completed,
            "ai_processing_minutes": usage.ai_processing_minutes,
            "api_calls": usage.api_calls
        }
    }


# ─────────────────────────────────────────────────────────────────────────────
# 21. GET /clinics/{clinic_id}/license — Monitoreo de licencias
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/{clinic_id}/license", summary="Monitoreo de utilización de licencias")
async def get_clinic_license(clinic_id: str = Path(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Clinic).where(Clinic.clinic_id == clinic_id))
    clinic = result.scalar_one_or_none()
    
    if not clinic:
        raise HTTPException(status_code=404, detail="Clínica no encontrada")

    used = getattr(clinic, 'patients_used', 0)
    limit = getattr(clinic, 'patients_limit', 0)
    utilization = round((used / limit) * 100, 2) if limit > 0 else 0

    return {
        "clinic_id": clinic_id,
        "licenses": {
            "total_limit": limit,
            "currently_active": used,
            "available": max(0, limit - used),
            "utilization_percentage": utilization
        },
        "warning_threshold_reached": utilization >= 90.0
    }


# ─────────────────────────────────────────────────────────────────────────────
# 22. GET /clinics/{clinic_id}/invoices — Historial de facturas
# FIX: Agregada compatibilidad de llaves 'id' y 'date' para que React no falle
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/{clinic_id}/invoices", summary="Historial de facturas emitidas")
async def get_clinic_invoices(clinic_id: str = Path(...), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Invoice)
        .where(Invoice.clinic_id == clinic_id)
        .order_by(desc(Invoice.issued_at))
    )
    invoices = result.scalars().all()

    pending_balance = sum(inv.amount for inv in invoices if getattr(inv, 'status', '') != "paid")

    return {
        "clinic_id": clinic_id,
        "pending_balance": pending_balance,
        "invoices": [
            {
                # React necesita estrictamente las llaves 'id' y 'date'. Aquí se las mandamos:
                "id": getattr(inv, 'invoice_id', getattr(inv, 'id', None)),
                "invoice_id": getattr(inv, 'invoice_id', getattr(inv, 'id', None)),
                "amount": inv.amount,
                "currency": getattr(inv, 'currency', 'USD'),
                "status": inv.status,
                "date": inv.issued_at.isoformat() if inv.issued_at else None,
                "issued_at": inv.issued_at.isoformat() if inv.issued_at else None,
                "pdf_url": getattr(inv, 'pdf_url', None)
            } for inv in invoices
        ]
    }


# ─────────────────────────────────────────────────────────────────────────────
# 23. POST /clinics/{clinic_id}/impersonate — Ingreso como soporte técnico
# ─────────────────────────────────────────────────────────────────────────────
@router.post("/{clinic_id}/impersonate", summary="Ingreso como soporte técnico", status_code=status.HTTP_201_CREATED)
async def impersonate_clinic(
    clinic_id: str = Path(...),
    body: dict = Body(...),
    db: AsyncSession = Depends(get_db)
):
    reason = body.get("reason", "")
    if len(reason) < 10:
        return {
            "success": False,
            "error": "La justificación ética debe tener más de 10 caracteres."
        }

    result = await db.execute(select(Clinic).where(Clinic.clinic_id == clinic_id))
    clinic = result.scalar_one_or_none()
    
    if not clinic:
        raise HTTPException(status_code=404, detail="Clínica no encontrada")

    return {
        "success": True,
        "message": "Impersonation session started successfully.",
        "session_id": f"sess_{uuid.uuid4().hex[:10]}",
        "temp_token": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9...",
        "expires_at": (datetime.utcnow()).isoformat(),
        "clinic_id": clinic_id,
        "reason_logged": reason
    }